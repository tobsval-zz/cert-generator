from openpyxl import load_workbook
from docx.shared import Pt
import translator
import docx
import glob

#Starter docx document search
starter_doc = glob.glob('./*.docx')[0].strip('.\\')
#Excel sheet setup
wb = load_workbook('excel_file.xlsx', data_only=True)
sheet = wb['']
cert_amount_range = (0, 0) #Change the amount here (see excel sheet, values in the tuple equal to the rows interval)

for i in range(*cert_amount_range):
    #individual_data >> certification id, name, grade, evaluation
    individual_data = (sheet.cell(row=i, column=1).value,
                       sheet.cell(row=i, column=2).value,
                       sheet.cell(row=i, column=20).value,
                       sheet.cell(row=i, column=21).value)
    #generates a new filename for each certificate (certificate owner's name + .docx extension)
    new_filename = str(individual_data[1].strip(' ')) + '.docx'

    #Doc and text settings setup
    doc = docx.Document(starter_doc)
    style = doc.styles['Normal']
    font = style.font
    font.name = 'Museo 300'
    font.size = Pt(10)

    #Substitute the name in the certificate's template with the owner's name
    for paragraph in doc.paragraphs:
        if '...' in paragraph.text:
            temp = paragraph.text
            paragraph.text = temp.replace('...', individual_data[1])
        
        #Substitute the score
        if '10 (dez valores)' in paragraph.text:
            temp = paragraph.text
            grade = str(individual_data[2])
            separated_grade = grade.split('.') #Separates floating point grades in two parts, which will then be used for translation
            if len(separated_grade) > 1:
                separated_grade[1] = separated_grade[1][:2] #Truncates the number of decimals in case of float scores (e.g.: 8.333, etc.)
            translated_grade = [translator.translate(elem) for elem in separated_grade]
            
            if len(translated_grade) > 1:
                formatted_grade = str(grade) + ' (' + translated_grade[0] + ' valores e ' + translated_grade[1] + ' decimais) '
                paragraph.text = temp.replace('10 (dez valores), ', formatted_grade)
            else:
                if translated_grade == ['']:
                    paragraph.text = temp.replace('10 (dez valores), ', 'não classificável ')
                else:
                    formatted_grade = str(grade) + ' (' + translated_grade[0] + ' valores) '
                    paragraph.text = temp.replace('10 (dez valores), ', formatted_grade)
        
        #Substitute the evaluation
        if 'Excelente' in paragraph.text:
            temp = paragraph.text
            paragraph.text = temp.replace('Excelente', individual_data[3])
               
        #Substitute the certificate's ID
        if '780' in paragraph.text:
            temp = paragraph.text
            paragraph.text = temp.replace('780', str(individual_data[0]))

    doc.save(new_filename)

    print(individual_data, ' -> Certificate Generated')
